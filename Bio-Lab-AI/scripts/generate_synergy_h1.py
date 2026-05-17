import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import random

random.seed(42)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Results"

ws.column_dimensions['A'].width = 14
for col in 'BCDEFGHIJKLM':
    ws.column_dimensions[col].width = 10

header_font = Font(bold=True)

ws['A1'] = "Synergy H1 Multi-Mode Reader"
ws['A1'].font = Font(bold=True, size=12)
ws['A2'] = ""
ws['A3'] = "Software Version: 3.12.08"
ws['A4'] = "Experiment Name: Cell Viability Assay - MTT"
ws['A5'] = "Date: 2026-05-02"
ws['A6'] = "Time: 14:32:05"
ws['A7'] = "Procedure: MTT Cell Viability"
ws['A8'] = "Wavelength: 570nm"
ws['A9'] = ""
ws['A10'] = "Results"
ws['A10'].font = header_font
ws['A11'] = ""

ws['A12'] = ""
for col_idx, col_num in enumerate(range(1, 13), start=2):
    cell = ws.cell(row=12, column=col_idx, value=col_num)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

row_labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

def r(lo, hi):
    return round(random.uniform(lo, hi), 3)

well_data = {
    'A': [r(0.85, 0.95) for _ in range(6)] + [r(0.40, 0.90) for _ in range(6)],
    'B': [r(0.45, 0.55) for _ in range(6)] + [r(0.40, 0.90) for _ in range(6)],
    'C': [r(0.70, 0.80) for _ in range(6)] + [r(0.40, 0.90) for _ in range(6)],
    'D': [r(0.20, 0.30) for _ in range(6)] + [r(0.40, 0.90) for _ in range(6)],
    'E': [r(0.05, 0.08) for _ in range(3)] + [r(0.40, 0.90) for _ in range(9)],
    'F': [0.00, 1.95] + [r(0.40, 0.90) for _ in range(10)],
    'G': [r(0.40, 0.90) for _ in range(12)],
    'H': [r(0.40, 0.90) for _ in range(12)],
}

for row_offset, row_label in enumerate(row_labels):
    excel_row = 13 + row_offset
    ws.cell(row=excel_row, column=1, value=row_label).font = header_font
    values = well_data[row_label]
    for col_offset, value in enumerate(values):
        cell = ws.cell(row=excel_row, column=2 + col_offset, value=value)
        cell.alignment = Alignment(horizontal='center')
        if value == 0.00 or value > 1.90:
            cell.font = Font(color="FF0000")

output_path = "test_synergy_h1_export.xlsx"
wb.save(output_path)
print(f"File created successfully: {output_path}")
